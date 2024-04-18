import tkinter as tk
from tkinter import messagebox, filedialog
import os, glob
import threading
from openpyxl import Workbook, load_workbook
import schedule
import time
from pystray import Icon as trayIcon, MenuItem as item, Menu as menu
from PIL import Image
import sys  # Import sys to use sys.exit()

# Declare tray_icon at the top of the script to ensure it's in the global scope
tray_icon = None

def select_folder(path_var):
    """Opens a dialog for the user to select a folder and updates the path variable."""
    folder_selected = filedialog.askdirectory()
    if folder_selected:
        path_var.set(folder_selected)

def check_and_prepare_files():
    global destination_path, source_path, keywords
    # Check destination file
    dest_file_path = os.path.join(destination_path.get(), "result.xlsx")
    if not os.path.isfile(dest_file_path):
        wb = Workbook()
        wb.save(filename=dest_file_path)

    # Check source path
    if not os.path.exists(source_path.get()):
        messagebox.showerror("Error", "Source folder path does not exist")
        return False

    return True

def process_files():
    if not check_and_prepare_files():
        return
    
    # Open destination Excel
    dest_file_path = os.path.join(destination_path.get(), "result.xlsx")
    wb_dest = load_workbook(dest_file_path)
    ws_dest = wb_dest.active

    # Ensure keywords are columns in the destination file
    keyword_list = [keyword.strip() for keyword in keywords.get().split(',')]
    
    # Identify the current headers
    current_headers = [cell.value for cell in ws_dest[1] if cell.value is not None]

    # Add new headers if they don't already exist
    column_to_add = len(current_headers) + 1
    for keyword in keyword_list:
        if keyword not in current_headers:
            ws_dest.cell(row=1, column=column_to_add, value=keyword)
            current_headers.append(keyword)
            column_to_add += 1
    
    # Save the workbook after adding headers
    wb_dest.save(dest_file_path)

    # Process each source file
    for file in glob.glob(f"{source_path.get()}/*.xlsx"):
        if file.endswith('result.xlsx'):
            continue

        wb_src = load_workbook(file, data_only=True)
        ws_src = wb_src.active

        # A dictionary to hold the first found value for each keyword
        found_values = {keyword: '' for keyword in keyword_list}
        
        for row in ws_src.iter_rows(min_row=2, values_only=True):
            for cell in row:
                if cell in keyword_list and found_values[cell] == '':
                    # Find the index of the keyword
                    keyword_index = row.index(cell)
                    # Assuming the value we need is right after the keyword
                    if keyword_index + 1 < len(row):
                        found_values[cell] = row[keyword_index + 1]

        # Append values to the destination workbook
        new_row = ws_dest.max_row + 1
        for keyword in keyword_list:
            col_index = current_headers.index(keyword) + 1
            ws_dest.cell(row=new_row, column=col_index, value=found_values[keyword])
    
    # Save the workbook after processing each file
    wb_dest.save(dest_file_path)
    messagebox.showinfo("Info", "Processing completed successfully!")

# Define a global running flag
running = True
tray_icon = None
scheduler_thread = None  # Keep a reference to the scheduler thread

# Modify the on_run function to use the running flag
def on_run():
    global running, scheduler_thread
    running = True
    scheduler_thread = threading.Thread(target=scheduled_process)
    scheduler_thread.start()

def scheduled_process():
    global running
    process_files()
    time_frame_val = int(time_frame.get())
    schedule.every(time_frame_val).minutes.do(process_files)

    while running:
        schedule.run_pending()
        time.sleep(1)

def on_stop():
    global running, tray_icon, scheduler_thread
    running = False  # Signal the thread to stop
    schedule.clear()  # Clear the schedule
    if tray_icon:
        tray_icon.stop()
    if scheduler_thread is not None:
        scheduler_thread.join()  # Wait for the thread to finish
    root.quit()  # Quit the Tkinter main loop
    root.destroy()  # This is necessary to destroy all widgets and close the window

def create_tray_icon():
    global tray_icon
    
    def exit_action(icon, item):
        on_stop()
    
    # Ensure you have an icon.png in the script directory
    icon_image = Image.open("icon.png")  
    menu_to_use = menu(item('Exit', exit_action))
    
    # Creating a new tray icon instance
    tray_icon = trayIcon("Test", icon_image, "My Application", menu_to_use)
    tray_icon.run()

# Setup the GUI
root = tk.Tk()
root.title("Excel Processor")

destination_path = tk.StringVar()
source_path = tk.StringVar()
keywords = tk.StringVar()
time_frame = tk.StringVar()

tk.Label(root, text="Destination Folder Path:").pack()
entry_dest = tk.Entry(root, textvariable=destination_path)
entry_dest.pack()
button_dest = tk.Button(root, text="Browse...", command=lambda: select_folder(destination_path))
button_dest.pack()

tk.Label(root, text="Source Folder Path:").pack()
entry_src = tk.Entry(root, textvariable=source_path)
entry_src.pack()
button_src = tk.Button(root, text="Browse...", command=lambda: select_folder(source_path))
button_src.pack()

tk.Label(root, text="Keywords (comma separated):").pack()
tk.Entry(root, textvariable=keywords).pack()
tk.Label(root, text="Time Frame (in minutes):").pack()
tk.Entry(root, textvariable=time_frame).pack()

tk.Button(root, text="Run", command=lambda: threading.Thread(target=on_run).start()).pack()
tk.Button(root, text="Stop", command=on_stop).pack()

root.protocol("WM_DELETE_WINDOW", create_tray_icon)
root.mainloop()
